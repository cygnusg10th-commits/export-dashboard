#!/usr/bin/env python3
"""
수출 통계 자동 업데이트 스크립트
공공데이터포털 관세청_품목별 수출입실적(GW) API 활용

사용법:
  python export_updater.py test    # API 연결 테스트
  python export_updater.py now     # 즉시 업데이트 실행
  python export_updater.py         # 스케줄러 모드 (매월 1·11·21일 자동)
"""

# ── 설정 (여기 3개만 수정하세요) ──────────────────────────────────────────────
API_KEY    = "발급받은_인증키_여기에_붙여넣기"
EXCEL_PATH = r"C:\수출입 통계\주요품목별수출정리_5월.xlsx"   # 최신 엑셀 파일 경로
RUN_DAYS   = (1, 11, 21)   # 매월 자동 실행 날짜 (잠정→갱신→확정)
# ─────────────────────────────────────────────────────────────────────────────

import re
import sys
import time
import logging
from datetime import date
from pathlib import Path

import requests
import openpyxl

# ── excel_editor 자동 탐색 ─────────────────────────────────────────────────
_here = Path(__file__).parent
for _d in [_here, _here / "dashboard", _here.parent / "dashboard"]:
    if (_d / "excel_editor.py").exists():
        sys.path.insert(0, str(_d))
        break

from excel_editor import add_monthly_row, add_quarter_row, find_last_data_row

# ── 로거 ───────────────────────────────────────────────────────────────────
_log_file = Path(__file__).parent / "export_updater.log"
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-7s  %(message)s",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler(_log_file, encoding="utf-8", mode="a"),
    ],
    force=True,
)
log = logging.getLogger(__name__)

BASE_URL = (
    "http://apis.data.go.kr"
    "/1220000/ItemExpImpInfoService1/getItemExpImpInfo1"
)


# ── 유틸 ───────────────────────────────────────────────────────────────────

def target_period() -> tuple[str, str]:
    """전월 반환: ('202504', '2025년04월')"""
    today = date.today()
    m, y = today.month - 1, today.year
    if m == 0:
        m, y = 12, y - 1
    return f"{y}{m:02d}", f"{y}년{m:02d}월"


def get_hs_code(ws) -> str | None:
    """시트 row 2, col 10 이후에서 HS코드(숫자 6자리 이상) 추출"""
    for c in range(10, 30):
        v = ws.cell(2, c).value
        if v:
            digits = re.sub(r"\D", "", str(v))
            if len(digits) >= 6:
                return digits
    return None


def _to_float(v) -> float | None:
    try:
        return float(str(v).replace(",", "").strip()) if v not in (None, "", "-") else None
    except Exception:
        return None


# ── API 호출 ───────────────────────────────────────────────────────────────

def fetch_export(hs_code: str, period_ym: str) -> dict | None:
    """
    HS코드 + 연월(YYYYMM)로 수출 실적 조회.
    Returns: {"dollar": float|None, "won": float|None, "weight": float|None} or None
    """
    hs_clean = re.sub(r"\D", "", str(hs_code))
    if len(hs_clean) < 6:
        return None
    hs_clean = hs_clean.ljust(10, "0")[:10]

    params = {
        "serviceKey": requests.utils.unquote(API_KEY),
        "hsCd":       hs_clean,
        "period":     period_ym,
        "type":       "json",
        "numOfRows":  "10",
        "pageNo":     "1",
    }
    try:
        resp = requests.get(BASE_URL, params=params, timeout=15)
        resp.raise_for_status()
        data = resp.json()
    except Exception as e:
        log.debug(f"요청 실패 HS{hs_clean[:6]}…: {e}")
        return None

    result_code = str(
        data.get("response", {})
            .get("header", {})
            .get("resultCode", "99")
    )
    if result_code not in ("00", "000", "0"):
        log.debug(f"비정상 응답 HS{hs_clean[:6]}…: code={result_code}")
        return None

    raw = (data.get("response", {})
               .get("body", {})
               .get("items"))
    if not raw:
        return None

    if isinstance(raw, dict):
        items = raw.get("item", [])
    else:
        items = raw
    if isinstance(items, dict):
        items = [items]
    if not items:
        return None

    item = items[0]
    return {
        "dollar": _to_float(item.get("expDlr")    or item.get("expDollar")),
        "won":    _to_float(item.get("expWon")     or item.get("expKrw")),
        "weight": _to_float(item.get("expWgt")     or item.get("expWeight")),
    }


# ── 메인 업데이트 ──────────────────────────────────────────────────────────

def run_update(excel_path: Path) -> None:
    """Excel 파일의 모든 시트를 API 데이터로 업데이트"""
    if not excel_path.exists():
        log.error(f"파일 없음: {excel_path}")
        return

    period_ym, period_label = target_period()
    log.info("=" * 60)
    log.info(f"업데이트 대상: {period_label}  ({excel_path.name})")

    wb = openpyxl.load_workbook(excel_path)
    sheets = wb.sheetnames[1:]   # 첫 번째 시트(목록) 제외

    ok = err = skip = already = 0

    for name in sheets:
        ws = wb[name]
        hs = get_hs_code(ws)
        if not hs:
            log.debug(f"[SKIP   ] {name:<30}  HS코드 없음")
            skip += 1
            continue

        last = find_last_data_row(ws)
        if last:
            last_val = str(ws.cell(last, 1).value or "")
            if period_label in last_val:
                log.debug(f"[EXISTS ] {name:<30}  {period_label} 이미 존재")
                already += 1
                continue

        data = fetch_export(hs, period_ym)
        if not data or data["dollar"] is None:
            log.warning(f"[NO DATA] {name:<28}  HS:{hs[:6]}… 데이터 없음")
            skip += 1
            continue

        try:
            new_row, is_qend = add_monthly_row(
                wb, name, period_label,
                dollar=data["dollar"],
                won=data["won"] or 0,
                weight=data["weight"] or 0,
            )
            if is_qend:
                add_quarter_row(wb, name, new_row)
            log.info(
                f"[OK    ] {name:<30}  "
                f"${data['dollar']:>15,.0f}  "
                f"{(data['weight'] or 0):>12,.0f} kg"
            )
            ok += 1
        except Exception as e:
            log.error(f"[ERR   ] {name}: {e}")
            err += 1

    log.info("-" * 60)
    log.info(f"결과: 업데이트={ok}  이미있음={already}  스킵={skip}  오류={err}")

    if ok:
        wb.save(excel_path)
        log.info(f"저장 완료: {excel_path}")
    log.info("=" * 60 + "\n")


# ── 테스트 ─────────────────────────────────────────────────────────────────

def run_test() -> None:
    """API 연결 테스트 — 반도체 HS코드(8542319100)로 샘플 조회"""
    period_ym, period_label = target_period()
    log.info(f"테스트: {period_label}  반도체 HS 8542319100 조회 중…")

    data = fetch_export("8542319100", period_ym)
    if data and data["dollar"] is not None:
        log.info("✅ API 연결 성공!")
        log.info(f"   수출 달러 : ${data['dollar']:,.0f}")
        log.info(f"   수출 중량 : {data['weight'] or 0:,.0f} kg")
        log.info("   → 이제 'python export_updater.py now' 로 업데이트하세요.")
    else:
        log.error("❌ API 연결 실패 또는 데이터 없음")
        log.error("   확인사항:")
        log.error("   1. API_KEY 값이 올바른지 확인")
        log.error(f"   2. {period_ym} 데이터가 아직 미공개일 수 있음")
        log.error("   3. data.go.kr 마이페이지 > 개발계정 > 서비스 승인 여부 확인")


# ── 스케줄러 ───────────────────────────────────────────────────────────────

def scheduled_job() -> None:
    today = date.today()
    if today.day in RUN_DAYS:
        log.info(f"스케줄 실행일: {today}")
        run_update(Path(EXCEL_PATH))


# ── 진입점 ─────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    cmd = sys.argv[1].lower() if len(sys.argv) > 1 else ""

    if cmd == "test":
        run_test()

    elif cmd == "now":
        run_update(Path(EXCEL_PATH))

    else:
        try:
            import schedule as _sched
        except ImportError:
            print("schedule 라이브러리가 없습니다: pip install schedule")
            sys.exit(1)

        log.info(f"스케줄러 시작 — 매일 09:00 체크, {RUN_DAYS}일에 업데이트 실행")
        log.info(f"엑셀 파일: {EXCEL_PATH}")

        scheduled_job()   # 시작 시 오늘이 실행일이면 즉시 1회 실행

        _sched.every().day.at("09:00").do(scheduled_job)
        while True:
            _sched.run_pending()
            time.sleep(60)
