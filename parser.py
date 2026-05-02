"""
수출입 통계 Excel → SQLite ETL
구조 확인:
  row1 col12: 기업명
  row2 col12: HS코드
  row4     : 헤더 (날짜, 분기수출금액, 분기단가, QoQ, YoY, 수출금액, MoM, YoY, 금액/중량, 가격YoY ...)
  row5~    : 데이터
  (0-indexed) col0=날짜, col5=일평균, col6=MoM, col7=YoY, col8=$/kg, col9=단가YoY
              col16=월총액($), col17=월총액(원), col18=중량(kg)
"""
import sqlite3, re, sys, logging
from pathlib import Path
from datetime import datetime

import openpyxl

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s", datefmt="%H:%M:%S")
logger = logging.getLogger(__name__)

SOURCE_DIR = Path(r"C:\수출입 통계")
DB_PATH    = Path(__file__).parent / "export_data.db"

# 0-based 컬럼 인덱스
C_DATE       = 0
C_AVG        = 5   # 일평균 수출액($)
C_MOM        = 6
C_YOY        = 7
C_UPRICE     = 8   # $/kg
C_PYOY       = 9   # 단가 YoY
C_DOLLAR     = 16  # 월간 총 수출액($)
C_WON        = 17  # 월간 총 수출액(₩)
C_WEIGHT     = 18  # 중량(kg)

MAX_COL = 30


def get_row(ws, rn: int) -> list:
    """read_only 모드에서 특정 행 값 리스트 반환"""
    for row in ws.iter_rows(min_row=rn, max_row=rn, min_col=1, max_col=MAX_COL, values_only=True):
        return list(row)
    return []


def parse_period(v) -> str | None:
    """'2024년04월' → '2024-04'  (한글 인코딩 무관하게 숫자만 추출)"""
    if not v:
        return None
    s = str(v)
    # 4자리 연도 + 비숫자 + 1-2자리 월 + 비숫자 패턴
    m = re.match(r"(\d{4})\D+(\d{1,2})\D*$", s.strip())
    if m:
        return f"{m.group(1)}-{int(m.group(2)):02d}"
    return None


def safe_f(v) -> float | None:
    if v is None:
        return None
    if isinstance(v, str) and (v.strip().startswith("#") or v.strip() == ""):
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def detect_dollar_weight_cols(ws, data_start: int) -> tuple[int, int]:
    """
    비표준 시트(미국·중국 등)의 달러/중량 컬럼을 동적 탐지.
    Strategy 1: 헤더행 텍스트에서 '달러'/'중량' 검색
    Strategy 2: dollar/weight ≈ unit_price 크로스체크 (unit_price 후보 자동 탐색)
    """
    # ── 전략 1: 헤더 텍스트 ──────────────────────────────────────────
    for header_rn in range(max(1, data_start - 3), data_start):
        col_dollar = col_weight = None
        for c1 in range(14, MAX_COL + 1):   # 1-indexed
            v = ws.cell(header_rn, c1).value
            if not isinstance(v, str):
                continue
            vs = v.strip()
            if "달러" in vs and "중량" not in vs and col_dollar is None:
                col_dollar = c1 - 1         # → 0-indexed
            elif ("중량" in vs or "kg" in vs.lower()) and col_weight is None:
                col_weight = c1 - 1
        if col_dollar is not None and col_weight is not None and col_dollar != col_weight:
            if (col_dollar, col_weight) != (C_DOLLAR, C_WEIGHT):
                logger.info("  헤더 기반 컬럼 탐지: dollar=%d weight=%d", col_dollar, col_weight)
            return (col_dollar, col_weight)

    # ── 전략 2: 크로스체크 ───────────────────────────────────────────
    sample_rows = []
    for rn in range(data_start, data_start + 15):
        r = get_row(ws, rn)
        if r and parse_period(r[C_DATE]):
            sample_rows.append(r)
        if len(sample_rows) >= 8:
            break

    # unit_price 후보: col 8 우선, 없으면 14~22 중 소값 컬럼 탐색
    up_col = None
    for test_col in [C_UPRICE] + list(range(14, min(23, MAX_COL))):
        ups = [safe_f(r[test_col]) for r in sample_rows
               if len(r) > test_col and safe_f(r[test_col]) is not None]
        if len(ups) >= 3 and all(0.05 < u < 5000 for u in ups):
            up_col = test_col
            break

    if up_col is None:
        return (C_DOLLAR, C_WEIGHT)

    candidates = range(14, min(23, MAX_COL))
    best = (C_DOLLAR, C_WEIGHT)
    best_score = 0
    for dc in candidates:
        for wc in candidates:
            if dc == wc or dc == up_col or wc == up_col:
                continue
            score = 0
            for r in sample_rows:
                up = safe_f(r[up_col]) if len(r) > up_col else None
                dv = safe_f(r[dc])     if len(r) > dc     else None
                wv = safe_f(r[wc])     if len(r) > wc     else None
                if up and dv and wv and wv != 0:
                    if abs(dv / wv - up) / up < 0.05:
                        score += 1
            if score > best_score:
                best_score = score
                best = (dc, wc)

    if best_score >= 3:
        if best != (C_DOLLAR, C_WEIGHT):
            logger.info("  크로스체크 기반 재매핑: dollar=%d weight=%d (score=%d)", best[0], best[1], best_score)
        return best
    return (C_DOLLAR, C_WEIGHT)


def parse_sheet(ws, sheet_name: str) -> tuple[dict, list[dict]]:
    # ── 메타 정보 ────────────────────────────────────────────────────
    # 기업명: row1, col10 이후 첫 번째 문자열
    company = None
    for c in range(10, MAX_COL + 1):
        v = ws.cell(1, c).value
        if v and isinstance(v, str) and v.strip():
            company = v.strip()
            break

    # HS코드: row2, col10 이후 첫 번째 숫자
    hs_code = None
    for c in range(10, MAX_COL + 1):
        v = ws.cell(2, c).value
        if v and isinstance(v, (int, float)) and float(v) > 0:
            hs_code = str(int(float(v)))
            break
        if v and isinstance(v, str) and re.match(r"^\d{4,}", v.strip()):
            hs_code = v.strip()
            break

    # ── 데이터 시작행 탐색 (한글 문자 비교 없이 숫자 패턴만 사용) ──
    data_start = None
    for rn in range(1, 12):
        v = ws.cell(rn, 1).value
        if v and parse_period(v):
            data_start = rn
            break

    if data_start is None:
        logger.warning("데이터행 없음: %s", sheet_name)
        return {}, []

    # ── 시트별 달러/중량 컬럼 동적 탐지 ────────────────────────────
    col_dollar, col_weight = detect_dollar_weight_cols(ws, data_start)

    # ── 데이터 파싱 ──────────────────────────────────────────────────
    records = []
    for rn in range(data_start, ws.max_row + 1):
        r = get_row(ws, rn)
        if not r:
            continue
        period = parse_period(r[C_DATE])
        if not period:
            continue

        def g(idx):
            return safe_f(r[idx]) if len(r) > idx else None

        records.append({
            "period":        period,
            "export_avg":    g(C_AVG),
            "mom":           g(C_MOM),
            "yoy":           g(C_YOY),
            "unit_price":    g(C_UPRICE),
            "price_yoy":     g(C_PYOY),
            "export_dollar": g(col_dollar),
            "export_won":    g(C_WON),
            "weight_kg":     g(col_weight),
        })

    info = {"sheet_name": sheet_name, "company": company, "hs_code": hs_code}
    return info, records


# ── DB ────────────────────────────────────────────────────────────────────────

SCHEMA = """
CREATE TABLE IF NOT EXISTS items (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    sheet_name  TEXT UNIQUE NOT NULL,
    company     TEXT,
    hs_code     TEXT,
    source_file TEXT,
    updated_at  TEXT
);
CREATE TABLE IF NOT EXISTS export_data (
    id            INTEGER PRIMARY KEY AUTOINCREMENT,
    sheet_name    TEXT NOT NULL,
    period        TEXT NOT NULL,
    export_avg    REAL,
    mom           REAL,
    yoy           REAL,
    unit_price    REAL,
    price_yoy     REAL,
    export_dollar REAL,
    export_won    REAL,
    weight_kg     REAL,
    UNIQUE(sheet_name, period)
);
CREATE INDEX IF NOT EXISTS idx_exp ON export_data(sheet_name, period);
CREATE TABLE IF NOT EXISTS parse_log (
    id            INTEGER PRIMARY KEY AUTOINCREMENT,
    file_path     TEXT,
    parsed_at     TEXT,
    sheets_count  INTEGER,
    records_count INTEGER
);
"""

def init_db(conn):
    conn.executescript(SCHEMA)
    conn.commit()


# ── ETL ───────────────────────────────────────────────────────────────────────

def parse_and_store(excel_path: Path, conn: sqlite3.Connection):
    logger.info("▶ 파싱: %s", excel_path.name)
    wb = openpyxl.load_workbook(str(excel_path), data_only=True, read_only=True)

    sheet_names = wb.sheetnames[1:]   # 첫 시트(목록·설정) 제외
    total_sheets = total_records = 0
    now = datetime.now().isoformat()

    for sheet_name in sheet_names:
        try:
            ws = wb[sheet_name]
            info, records = parse_sheet(ws, sheet_name)
            if not records:
                continue

            conn.execute("""
                INSERT INTO items (sheet_name, company, hs_code, source_file, updated_at)
                VALUES (?,?,?,?,?)
                ON CONFLICT(sheet_name) DO UPDATE SET
                    company=excluded.company, hs_code=excluded.hs_code,
                    source_file=excluded.source_file, updated_at=excluded.updated_at
            """, (sheet_name, info.get("company"), info.get("hs_code"), excel_path.name, now))

            conn.executemany("""
                INSERT INTO export_data
                    (sheet_name,period,export_avg,mom,yoy,unit_price,price_yoy,
                     export_dollar,export_won,weight_kg)
                VALUES (?,?,?,?,?,?,?,?,?,?)
                ON CONFLICT(sheet_name,period) DO UPDATE SET
                    export_avg=excluded.export_avg, mom=excluded.mom, yoy=excluded.yoy,
                    unit_price=excluded.unit_price, price_yoy=excluded.price_yoy,
                    export_dollar=excluded.export_dollar, export_won=excluded.export_won,
                    weight_kg=excluded.weight_kg
            """, [(sheet_name, r["period"], r["export_avg"], r["mom"], r["yoy"],
                   r["unit_price"], r["price_yoy"], r["export_dollar"],
                   r["export_won"], r["weight_kg"]) for r in records])

            conn.commit()
            total_sheets  += 1
            total_records += len(records)

        except Exception as exc:
            logger.error("시트 오류 [%s]: %s", sheet_name, exc)
            conn.rollback()

    wb.close()
    conn.execute(
        "INSERT INTO parse_log(file_path,parsed_at,sheets_count,records_count) VALUES(?,?,?,?)",
        (str(excel_path), now, total_sheets, total_records),
    )
    conn.commit()
    logger.info("✓ 완료: %d 품목, %d 건", total_sheets, total_records)
    return total_sheets, total_records


def find_excel_files() -> list[Path]:
    return sorted(SOURCE_DIR.glob("주요품목별수출정리*.xlsx"))


def needs_reparse(excel_path: Path, conn: sqlite3.Connection) -> bool:
    row = conn.execute(
        "SELECT parsed_at FROM parse_log WHERE file_path=? ORDER BY parsed_at DESC LIMIT 1",
        (str(excel_path),),
    ).fetchone()
    if not row:
        return True
    return datetime.fromtimestamp(excel_path.stat().st_mtime).isoformat() > row[0]


def run(force: bool = False):
    DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(str(DB_PATH))
    init_db(conn)

    files = find_excel_files()
    if not files:
        logger.warning("Excel 파일 없음: %s", SOURCE_DIR)
        conn.close()
        return

    for f in files:
        if force or needs_reparse(f, conn):
            parse_and_store(f, conn)
        else:
            logger.info("최신 상태 (스킵): %s", f.name)

    conn.close()


if __name__ == "__main__":
    run(force="--force" in sys.argv)
