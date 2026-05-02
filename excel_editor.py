"""
수출 통계 Excel 편집 유틸리티
- 기존 시트에 신규 월 데이터 행 추가 (수식 패턴 자동 연장)
- 새 품목 시트 추가 (기존 시트 구조 복제)
"""
import re
import io
from copy import copy
from typing import Optional

import openpyxl
import pandas as pd


# ── 수식 처리 ─────────────────────────────────────────────────────────────────

def shift_row_refs(formula: str, delta: int) -> str:
    """Excel 수식 내 상대 행 참조를 delta만큼 이동 (절대 참조 $ 유지)"""
    if not isinstance(formula, str) or not formula.startswith("="):
        return formula

    def replace(m):
        abs_col, col, abs_row, row = m.group(1), m.group(2), m.group(3), m.group(4)
        if abs_row == "$":
            return f"{abs_col}{col}${row}"
        return f"{abs_col}{col}{int(row) + delta}"

    return "=" + re.sub(r"(\$?)([A-Za-z]+)(\$?)(\d+)", replace, formula[1:])


# ── 행·컬럼 탐색 ──────────────────────────────────────────────────────────────

def find_last_data_row(ws) -> Optional[int]:
    """날짜 패턴(\d{4})이 있는 마지막 행 번호"""
    last = None
    for row in ws.iter_rows(min_col=1, max_col=1):
        v = row[0].value
        if v and re.match(r"\d{4}", str(v).strip()):
            last = row[0].row
    return last


def find_first_data_row(ws) -> Optional[int]:
    """날짜 패턴이 있는 첫 번째 행 번호"""
    for row in ws.iter_rows(min_col=1, max_col=1):
        v = row[0].value
        if v and re.match(r"\d{4}", str(v).strip()):
            return row[0].row
    return None


def find_last_quarter_row(ws, before_row: int) -> Optional[int]:
    """before_row 이전 가장 최근 분기 집계 행 (col A 비어 있고 col B 수식)"""
    for rn in range(before_row - 1, 0, -1):
        a = ws.cell(rn, 1).value
        b = ws.cell(rn, 2).value
        if (a is None or str(a).strip() == "") and isinstance(b, str) and b.startswith("="):
            return rn
    return None


def detect_input_columns(ws, near_row: int) -> dict:
    """
    헤더행 스캔으로 달러/원화/중량/영업일수 컬럼 위치(1-indexed) 탐지.
    못 찾으면 parser.py 기본값(C_DOLLAR=16→col17 등) 사용.
    """
    mapping: dict[str, Optional[int]] = {}
    kw_map = {
        "dollar":   ["금액(달러)", "달러", "USD", "수출금액"],
        "won":      ["금액(원화)", "원화", "KRW"],
        "weight":   ["중량", "KG", "Kg"],
        "workdays": ["영업일수", "영업일", "일수"],
    }
    for rn in range(max(1, near_row - 6), near_row):
        for cell in ws[rn]:
            v = str(cell.value or "").strip()
            for key, kws in kw_map.items():
                if key not in mapping and any(k in v for k in kws):
                    mapping[key] = cell.column
    return {
        "dollar":   mapping.get("dollar",   17),
        "won":      mapping.get("won",       18),
        "weight":   mapping.get("weight",   19),
        "workdays": mapping.get("workdays"),
    }


# ── 셀 행 복사 ────────────────────────────────────────────────────────────────

def clone_row(ws, from_row: int, to_row: int, delta: int) -> None:
    """from_row의 값·수식·스타일을 to_row에 복사 (수식 행번호 delta 이동)"""
    for col_idx in range(1, ws.max_column + 1):
        src = ws.cell(from_row, col_idx)
        dst = ws.cell(to_row, col_idx)
        if src.has_style:
            dst.font          = copy(src.font)
            dst.fill          = copy(src.fill)
            dst.border        = copy(src.border)
            dst.alignment     = copy(src.alignment)
            dst.number_format = src.number_format
        v = src.value
        dst.value = shift_row_refs(v, delta) if isinstance(v, str) and v.startswith("=") else v


# ── 핵심: 신규 월 행 추가 ─────────────────────────────────────────────────────

def add_monthly_row(
    wb: openpyxl.Workbook,
    sheet_name: str,
    period: str,        # "2025년05월"
    dollar: float,
    won: float,
    weight: float,
    workdays: Optional[int] = None,
) -> tuple[int, bool]:
    """
    시트에 신규 월 데이터 행 추가 (마지막 행 수식 패턴 연장).
    Returns: (new_row, is_quarter_end)
    """
    ws = wb[sheet_name]

    last_row = find_last_data_row(ws)
    if last_row is None:
        raise ValueError(f"'{sheet_name}' 시트에서 데이터 행을 찾을 수 없습니다")

    new_row = last_row + 1
    cols = detect_input_columns(ws, last_row)

    clone_row(ws, last_row, new_row, delta=1)

    ws.cell(new_row, 1).value               = period
    ws.cell(new_row, cols["dollar"]).value  = dollar
    ws.cell(new_row, cols["won"]).value     = won
    ws.cell(new_row, cols["weight"]).value  = weight
    if workdays and cols.get("workdays"):
        ws.cell(new_row, cols["workdays"]).value = workdays

    m = re.search(r"(\d{1,2})월", period)
    is_qend = m and int(m.group(1)) in (3, 6, 9, 12)
    return new_row, bool(is_qend)


def add_quarter_row(
    wb: openpyxl.Workbook,
    sheet_name: str,
    after_data_row: int,
) -> int:
    """분기 집계 행을 after_data_row 다음에 추가. Returns: 추가된 행 번호."""
    ws = wb[sheet_name]
    q_row = after_data_row + 1

    prev_q = find_last_quarter_row(ws, after_data_row)
    if prev_q is None:
        return after_data_row

    clone_row(ws, prev_q, q_row, delta=(q_row - prev_q))
    ws.cell(q_row, 1).value = None   # 분기 행 col A 는 비움
    return q_row


# ── KITA / 관세청 파일 파싱 ───────────────────────────────────────────────────

def parse_kita_file(file_bytes: bytes, filename: str) -> Optional[dict]:
    """
    관세청/KITA 수출 데이터 파일 파싱.
    Returns dict: {period, dollar, won, weight, workdays, raw_df} 또는 None
    """
    try:
        if filename.lower().endswith(".csv"):
            df = pd.read_csv(io.BytesIO(file_bytes), encoding="utf-8", thousands=",")
        else:
            # 헤더 자동 탐지 (첫 10행 중 한글 컬럼명 있는 행)
            df = None
            for hrow in range(10):
                tmp = pd.read_excel(io.BytesIO(file_bytes), header=hrow)
                if any("금액" in str(c) or "중량" in str(c) or "년월" in str(c)
                       for c in tmp.columns):
                    df = tmp
                    break
            if df is None:
                df = pd.read_excel(io.BytesIO(file_bytes))
    except Exception:
        return None

    df.columns = [str(c).strip() for c in df.columns]
    col_lower  = {c.lower(): c for c in df.columns}

    def find_col(*kws):
        for kw in kws:
            for k, orig in col_lower.items():
                if kw.lower() in k:
                    return orig
        return None

    dc = find_col("달러", "usd", "금액(달러)", "수출금액")
    wc = find_col("원화", "krw", "금액(원화)")
    kc = find_col("중량", "kg")
    tc = find_col("년월", "기간", "월")
    ec = find_col("영업일")

    def to_float(v):
        try:
            return float(str(v).replace(",", "").strip())
        except Exception:
            return None

    for _, row in df.iterrows():
        period_raw = str(row.get(tc, "") if tc else "").strip()
        m = re.search(r"(\d{4})[\s년/-]?0*(\d{1,2})", period_raw)
        if not m:
            continue
        period = f"{m.group(1)}년{int(m.group(2)):02d}월"
        dollar   = to_float(row.get(dc))   if dc else None
        won      = to_float(row.get(wc))   if wc else None
        weight   = to_float(row.get(kc))   if kc else None
        workdays = int(float(row.get(ec))) if ec and row.get(ec) else None
        if dollar is None and won is None:
            continue
        return {
            "period": period, "dollar": dollar, "won": won,
            "weight": weight, "workdays": workdays,
            "detected_cols": {"dollar": dc, "won": wc, "weight": kc,
                              "date": tc, "workdays": ec},
            "raw_df": df,
        }
    return None


# ── 새 시트 추가 ──────────────────────────────────────────────────────────────

def add_new_sheet(
    wb: openpyxl.Workbook,
    new_name: str,
    template_name: str,
    company: str,
    hs_code: str,
    data_rows: list[dict],   # [{period, dollar, won, weight, workdays}, ...]
) -> None:
    """
    template_name 시트를 복사해 new_name 시트 생성 후 데이터 채움.
    data_rows는 기간 오름차순 정렬된 리스트.
    """
    src_ws = wb[template_name]
    new_ws = wb.copy_worksheet(src_ws)
    new_ws.title = new_name

    # 메타 정보 교체 (기업명 row1, HS코드 row2)
    for rn, val in [(1, company), (2, hs_code)]:
        for c in range(10, 30):
            v = new_ws.cell(rn, c).value
            if v is not None and str(v).strip():
                new_ws.cell(rn, c).value = val
                break

    first_row = find_first_data_row(new_ws)
    last_row  = find_last_data_row(new_ws)
    if first_row is None:
        return

    cols = detect_input_columns(new_ws, first_row)

    # 기존 데이터 행 값 초기화
    for rn in range(first_row, (last_row or first_row) + 1):
        a = new_ws.cell(rn, 1).value
        if a and re.match(r"\d{4}", str(a).strip()):
            for ci in range(1, new_ws.max_column + 1):
                new_ws.cell(rn, ci).value = None

    # 신규 데이터 채우기
    template_data_row = first_row   # 수식 패턴 참조용 원본 행
    for i, rd in enumerate(data_rows):
        target_row = first_row + i
        if i > 0:
            clone_row(new_ws, target_row - 1, target_row, delta=1)
        new_ws.cell(target_row, 1).value              = rd["period"]
        if rd.get("dollar") is not None:
            new_ws.cell(target_row, cols["dollar"]).value = rd["dollar"]
        if rd.get("won") is not None:
            new_ws.cell(target_row, cols["won"]).value    = rd["won"]
        if rd.get("weight") is not None:
            new_ws.cell(target_row, cols["weight"]).value = rd["weight"]
        if rd.get("workdays") and cols.get("workdays"):
            new_ws.cell(target_row, cols["workdays"]).value = rd["workdays"]
